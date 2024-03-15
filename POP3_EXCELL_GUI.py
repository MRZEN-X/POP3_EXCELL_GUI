import atexit
import shutil
import sys
import poplib
import email
from email.header import decode_header
import os
import threading
from PyQt5.QtWidgets import QApplication, QWidget, QVBoxLayout, QLabel, QLineEdit, QPushButton, QTextEdit,QHBoxLayout
from PyQt5.QtCore import pyqtSignal
from datetime import datetime, timedelta
import locale
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import warnings
warnings.filterwarnings("ignore")# !!!
import pandas as pd
import qdarkstyle
import subprocess

class EmailLoginApp(QWidget):
    # 子线程添加通知信号
    update_output_text_signal = pyqtSignal(str)

    def __init__(self):
        super().__init__()
        self.logged_in = False
        self.email_address = None
        self.password = None
        self.server_address = None
        self.email_address = None
        self.mail = None
        self.email_count = int(1)
        # Unix LIke SYS 打包获取路径  os.getcwd()失效
        # if getattr(sys, 'frozen', False):
        #     self.now_path = os.path.dirname(sys.executable)
        # elif __file__:
        #     self.now_path = os.path.dirname(__file__)
        
        # 固定路径
        self.now_path="????待定！！！！！！！！！！！！！！"

        # windows和unix类直接运行python文件时可用
        # self.now_path = os.getcwd()

        self.init_ui()
        # 绑定信号到self.update_output_text_in_thread
        self.update_output_text_signal.connect(self.update_output_text_in_thread)

    # 基本布局
    def init_ui(self):
        self.setWindowTitle("～～xxxxxxxxxxxx拉取、汇总～～")
        self.setGeometry(100, 100, 600, 500)
        # 定义一堆组建
        self.email_label = QLabel("邮箱地址:")
        self.email_input = QLineEdit()
        self.email_input.setText("xxxxxxxxxxxxxxx")
        self.password_label = QLabel("密码:")
        self.password_input = QLineEdit()
        #self.password_input.setPlaceholderText("  密码不可见o.O ")
        self.password_input.setText("xxxxxxxxxxxxxxxxxxxx")
        self.password_input.setEchoMode(QLineEdit.Password)
        self.server_label = QLabel("邮件服务器地址:")
        self.server_input = QLineEdit()
        self.server_input.setText("xxxxxxxxxxxxxxxxxxxxxxxxx")
        self.login_button = QPushButton("登录")
        self.login_button.clicked.connect(self.login_to_email)
        # 全局输出框
        self.output_text = QTextEdit()
        self.output_text.setReadOnly(True)

        # 登陆成功后显示的功能界面
        self.date_label = QLabel("～～今日日期："+datetime.now().strftime('%Y.%m.%d')+" ～～\n向前拉取邮件范围:0-0今  0-1今、昨  1-3昨、前、大前  3-9....\n")
        self.date_input = QLineEdit()
        self.date_input.setPlaceholderText("请输入拉取邮件范围")
        self.process_button = QPushButton("获取邮件：一次获取完成后再执行其他任务 防止多线程拉邮件被服务器封ip")
        self.process_button.clicked.connect(self.process_emails_threaded)
        # 添加一个按钮来打开self.now_path文件夹
        self.open_folder_button = QPushButton("打开文件夹")
        self.open_folder_button.clicked.connect(self.open_folder)

        # 添加布局
        layout = QVBoxLayout()
        
        layout.addWidget(self.email_label)
        layout.addWidget(self.email_input)
        layout.addWidget(self.password_label)
        layout.addWidget(self.password_input)
        layout.addWidget(self.server_label)
        layout.addWidget(self.server_input)
        layout.addWidget(self.login_button)
        layout.addWidget(self.output_text)
        layout.addWidget(self.date_label)
        layout.addWidget(self.date_input)
        layout.addWidget(self.process_button)
        layout.addWidget(self.open_folder_button)

         # 登陆成功后显示的部件，登陆界面不可见
        self.date_label.setVisible(False)
        self.date_input.setVisible(False)
        self.process_button.setVisible(False)
        self.open_folder_button.setVisible(False)

        self.setLayout(layout)


    # 槽函数，用于在主线程中执行添加提示框内容的操作
    def update_output_text_in_thread(self, text):
        self.output_text.append(text)
    
    # 登陆函数
    def login_to_email(self):
        self.email_address = self.email_input.text()
        self.password = self.password_input.text()
        self.server_address = self.server_input.text()

        try:
            # 连接到邮箱服务器
            self.mail = poplib.POP3_SSL(self.server_address)

            # 登录到邮箱
            self.mail.user(self.email_address)
            self.mail.pass_(self.password)

            # 获取邮箱中的邮件数量
            self.email_count = len(self.mail.list()[1])
            self.output_text.clear()
            self.output_text.append(f"登录成功\n服务器中共有邮件：{self.email_count}")

            # 登陆标记
            self.logged_in = True

            # 登陆后登陆界面不可见
            self.email_label.setVisible(False)
            self.email_input.setVisible(False)
            self.password_label.setVisible(False)
            self.password_input.setVisible(False)
            self.server_label.setVisible(False)
            self.server_input.setVisible(False)
            self.login_button.setVisible(False)

            # 登陆成功，输出登陆结果，改变布局隐藏部分组件
            self.date_label.setVisible(True)
            self.date_input.setVisible(True)
            self.process_button.setVisible(True)
            self.open_folder_button.setVisible(True)

        except Exception as e:
            self.output_text.append(f"登录失败: {str(e)}")
        
    # 获取拉取天数范围，使用子线程拉取文件避免主线程界面卡顿
    def process_emails_threaded(self):
        if not self.logged_in:
            return

        try:
            resp, lines, octets = self.mail.retr(self.email_count-1)
        except poplib.error_proto as e:
            self.mail = poplib.POP3_SSL(self.server_address)
            self.mail.user(self.email_address)
            self.mail.pass_(self.password)
            self.output_text.append(f"连接断开！ 已重连 : {str(e)}")

    
        # 坑中坑 草泥马   ！！！！！！！！！！！！！！！！！！！！！！！！！！！！
        locale.setlocale(locale.LC_ALL,'en_US.UTF-8')
        # 解析输入为日期范围
        input_str = self.date_input.text()
        values = input_str.split('-')

        if len(values) != 2:
            self.output_text.append("未能解析天数范围，请重新输入。")
            return

        num1 = int(values[0])
        num2 = int(values[1])

        if num1 > num2:
            self.output_text.append("num1必须小于等于num2，请重新输入!")
            return

        days_ago1 = (datetime.now() - timedelta(days=num1)).date() # 近端
        days_ago2 = (datetime.now() - timedelta(days=num2)).date() # 远端
        
        # 显示拉取范围
        if days_ago1 == days_ago2:
            self.output_text.append(f"正在读取：{days_ago2.strftime('%Y.%m.%d')} 的邮件,请等候～～")
        else:
            self.output_text.append(f"正在读取：{days_ago2.strftime('%Y.%m.%d')} -- {days_ago1.strftime('%Y.%m.%d')} 的邮件,请等候～～")


        # 调起调动子线程
        try:
            email_thread = threading.Thread(target=self.get_emails, args=( days_ago1, days_ago2))
            email_thread.start()
        except Exception as e:
            print(f"获取失败: {str(e)}")

         
                 
    # 下载并保存附件  
    def get_emails(self, days_ago1, days_ago2):       
        for i in range(self.email_count ,0, -1):
            # 获取邮件内容
            resp, lines, octets = self.mail.retr(i)
            msg_data = b'\n'.join(lines)
            msg = email.message_from_bytes(msg_data)

            # 获取邮件日期
            date_str = msg["Date"]
            msg_date = datetime.strptime(date_str, "%a, %d %b %Y %H:%M:%S %z").date()

            #优化渐进日期算法
            if (msg_date - days_ago1).days > 1:
                i = i-8*(msg_date-days_ago1).days+8
            if days_ago2 > msg_date:
                break
            
            # 判断发件人
            from_email = msg.get("From", "")
            if "发件人的邮箱地址  xxxxx" not in from_email:
                continue
            

            # 判断是否在拉取日期范围
            if msg_date >= days_ago2 and msg_date <= days_ago1:
                # 获取邮件主题
                subject, encoding = decode_header(msg["Subject"])[0]
                if isinstance(subject, bytes):
                    try:
                        subject = subject.decode(encoding if encoding else "utf-8")
                    except Exception as e:
                        subject = subject.decode("gb18030")
                        encoding = "gb18030"

                # 处理附件
                for part in msg.walk():
                    if part.get_content_maintype() == "multipart":
                        continue
                    if part.get("Content-Disposition") is None:
                        continue

                    filename = part.get_filename()
                    if filename:
                        filename = decode_header(filename)[0][0]
                        if isinstance(filename, bytes):
                            # 编码是个大坑
                            try:
                                filename = filename.decode(encoding if encoding else "utf-8")
                            except Exception as e:
                                pass
                            #文件名判断
                            if  "附件中需要的存在字段： xxxxxxxxxxxxx" in filename:
                                # 使用信号向主线程发送消息
                                self.update_output_text_signal.emit(f"拉取："+filename+"   "+date_str)
                                filename = os.path.join(self.now_path, filename)
                                # 保存附件到当前目录
                                with open(filename, "wb") as attachment_file:
                                    attachment_file.write(part.get_payload(decode=True))
        
        # 输出获取结果
        if days_ago1 == days_ago2:
            self.update_output_text_signal.emit(f"{days_ago2.strftime('%Y.%m.%d')} 获取完成")
        else:
            self.update_output_text_signal.emit(f"{days_ago2.strftime('%Y.%m.%d')} -- {days_ago1.strftime('%Y.%m.%d')}  获取完成")

        #///////////////////////////////////////////////////汇总数据///////////////////////////////////////////////
        # 获取当前目录下的所有Excel文件
        excel_files = [file for file in os.listdir(self.now_path) if file.endswith('.xlsx') and "附件中需要的存在字段： xxxxxxxxxxxxx" in file]
        # 创建一个空的DataFrame来存储数据
        all_data = pd.DataFrame()
        # 存储文件名称
        data_sources = []
        # 循环遍历每个Excel文件并提取第一列数据，将它们合并到all_data的第一列中
        for file in excel_files:
            # 备注文件名获取
            parts = file.split('-')
            data_sources_name = parts[1]+"_"+parts[2][:-5]+"_XWSOC"

            df = pd.read_excel(os.path.join(self.now_path, file), header=0)  # 忽略每列的标题
            first_column = df.iloc[:, 0]
            # 添加文件名数据
            for i in range(len(first_column)):
                data_sources.append(data_sources_name)

            all_data = pd.concat([all_data, first_column], axis=0, ignore_index=True)

        # 保存数据到1.csv
        # all_data.to_csv('1.csv', header=False, index=False)

        #打开汇总.xlsx
        sum_cell = load_workbook(os.path.join(self.now_path, '汇总.xlsx'))
        sheet = sum_cell.active
        # 将数据从 DataFrame 转换为行列表
        data_rows = list(dataframe_to_rows(all_data, index=False, header=False))
    
        #检测冲突
        for i in range(2, 6000):
            cell_value = sheet.cell(row=i, column=6).value
            if cell_value:
                sheet.cell(row=i, column=6,value = '')
            cell_value = sheet.cell(row=i, column=8).value
            if cell_value:
                sheet.cell(row=i, column=8,value = '')
        self.update_output_text_signal.emit(f"清空汇总表")
            
        # 将数据写入现有文件的第五列，从第二行开始
        for i, row in enumerate(data_rows, start=2):
            #print(i)
            for j, value in enumerate(row, start=6):
                #print(j)
                sheet.cell(row=i, column=j, value=value)
                sheet.cell(row=i, column=8, value=data_sources[i-2]) # 匹配错位数
                

        # 保存修改后的数据到"汇总.xlsx"文件
        sum_cell.save(os.path.join(self.now_path, '汇总.xlsx'))

        #///////////////////////////////////////////////////按日期转移拉取附件////////////////////////////////////////////////
        # 创建以当前日期为名称的文件夹
        if days_ago1 == days_ago2:
            folder_name = days_ago2.strftime('%Y.%m.%d')
        else:
            folder_name = days_ago2.strftime('%Y.%m.%d')+" -- "+ days_ago1.strftime('%Y.%m.%d')
        os.makedirs(os.path.join(self.now_path,folder_name), exist_ok=True)
        # 获取当前目录下的所有文件
        files = [file for file in os.listdir(self.now_path) if "附件中需要的存在字段： xxxxxxxxxxxxx" in file]
        # 移动所有包含“xxxxxxxxx”的文件到新创建的文件夹
        for file in files:
            source_path = os.path.join(self.now_path, file)
            destination_path = os.path.join(self.now_path, folder_name, file)
            shutil.move(source_path, destination_path)
        # 复制生成的汇总文件到生成的文件夹下
        shutil.copy(os.path.join(self.now_path, '汇总.xlsx'),os.path.join(self.now_path,folder_name))
        self.update_output_text_signal.emit(f"拉取文件已移动到文件夹： "+ folder_name + "\n")
        # //////////////////////////////拉取完成/////////////////////////////

    def open_folder(self):
        #subprocess.Popen(f'explorer {self.now_path}', shell=True)  # Windows中打开文件夹
        subprocess.Popen(['open', self.now_path])  # 在Mac中打开文件夹

    # 断开链接
    def quit_email(self):
        try:
            self.mail.quit()
            print("邮箱断开链接")
        except Exception as e:
            print("没登陆，断不开")

if __name__ == '__main__':
    #邮件数据长度扩展
    poplib._MAXLINE=20480
    app = QApplication(sys.argv)
    app.setStyleSheet(qdarkstyle.load_stylesheet(qt_api='pyqt5'))
    email_app = EmailLoginApp()
    email_app.show()
    # 链接quit_email(self):点击关闭窗口断开邮箱SSL链接
    atexit.register(email_app.quit_email)
    sys.exit(app.exec_())
